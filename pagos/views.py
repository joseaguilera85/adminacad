from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Prefetch
from .forms import PaymentPlanForm, PriceListForm
from .models import (
    Apartment,
    PaymentRecord,
    ListaCliente,
    PriceList,
    Project,
    PaymentTransaction,
    AccountStatement,
    Venta
)
from apartments.models import Apartment
from clientes.models import Cliente
from datetime import datetime, timedelta
from decimal import Decimal
import calendar
from django.utils import timezone

# -------------------------------

def menu(request):
    is_cliente = request.user.groups.filter(name="Cliente").exists()
    return render(request, "pagos/admin_venta_menu.html", {"is_cliente": is_cliente})

# -------------------------------
#### 3.1 Lista de precios ###
# -------------------------------

def price_list_view(request):
    # Get all projects for the dropdown
    projects = Project.objects.all()
    
    # Get the selected project ID from the GET request
    selected_project_id = request.GET.get('project')
    
    # Get the sorting field from the GET request
    order_by = request.GET.get('order_by', 'number')  # Default to 'number'

    # Filter apartments based on the selected project and apply ordering
    if selected_project_id:
        apartments = Apartment.objects.filter(project_id=selected_project_id).order_by(order_by)
    else:
        apartments = Apartment.objects.all().order_by(order_by)

    # Get price lists if needed
    price_lists = PriceList.objects.filter(apartment__in=apartments)

    context = {
        'projects': projects,
        'selected_project_id': selected_project_id,
        'apartments': apartments,
        'price_lists': price_lists,
        'order_by': order_by,  # Pass the current ordering to the template
    }
    return render(request, 'pagos/price_list.html', context)

#------------------------------

def edit_price_list_values(request, apartment_id):
    price_list, created = PriceList.objects.get_or_create(apartment_id=apartment_id)

    if request.method == 'POST':
        form = PriceListForm(request.POST, instance=price_list)
        if form.is_valid():
            form.save()
            return redirect('pagos:price_list')  # Redirect to the price list view
    else:
        form = PriceListForm(instance=price_list)

    context = {
        'form': form,
        'price_list': price_list,
    }
    return render(request, 'pagos/edit_price_list_values.html', context)

#------------------------------

def edit_current_price_list_index(request, apartment_id):
    price_list = get_object_or_404(PriceList, apartment_id=apartment_id)

    if request.method == 'POST':
        price_list.current_list_price_index = request.POST.get('current_list_price_index')
        price_list.save()
        return redirect('pagos:price_list')  # Redirect to the price list view

    context = {
        'price_list': price_list,
    }
    return render(request, 'pagos/edit_current_price_list_index.html', context)


#------------------------------
#### 3.2a Inventario ####
#------------------------------

from django.db.models import Prefetch

def inventario_view(request):
    # Get filter values from query parameters
    project_name_filter = request.GET.get('project_name', '')
    status_filter = request.GET.get('status', '')
    order_by = request.GET.get('order_by', 'number')  # Default sorting field

    # Get all projects to populate the dropdown
    all_projects = Project.objects.all()

    # Get all status choices from the Apartment model
    status_choices = Apartment.STATUS_CHOICES

    # Prepare the Prefetch object with filters applied
    apartments_queryset = Apartment.objects.order_by(order_by)
    if status_filter:
        apartments_queryset = apartments_queryset.filter(status=status_filter)

    apartments_prefetch = Prefetch('apartments', queryset=apartments_queryset)

    # Filter projects based on project_name_filter and prefetch apartments
    if project_name_filter:
        projects = Project.objects.prefetch_related(apartments_prefetch).filter(name=project_name_filter)
    else:
        projects = Project.objects.prefetch_related(apartments_prefetch)

    return render(request, 'pagos/inventario.html', {
        'projects': projects,
        'all_projects': all_projects,  # Pass all project names for the dropdown
        'status_choices': status_choices,  # Pass status choices for the dropdown
        'project_name_filter': project_name_filter,  # Pass the current project filter
        'status_filter': status_filter,  # Pass the current status filter
        'order_by': order_by,  # Pass the sorting field
    })




#------------------------------
#### 3.2b Vista edificio / Plano ####
# -----------------------------

from django.shortcuts import render
from .models import Project, Apartment

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.core.serializers import serialize

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
import json

def plano_lotes(request):
    # Get all projects for the dropdown
    projects = Project.objects.all()

    # Get the selected project ID from the GET request
    selected_project_id = request.GET.get('project')

    # Initialize variables
    apartments = []
    houses_json = None
    project_name = "All Projects"

    if selected_project_id:
        # Validate and filter apartments for the selected project
        selected_project = get_object_or_404(Project, id=selected_project_id)
        project_name = selected_project.name
        apartments = Apartment.objects.filter(project=selected_project).order_by('number')

    # Create a list of apartments with necessary properties
    apartments_with_properties = [
        {
            "id": apartment.number,
            "status": apartment.status,
            "points": apartment.points,  # Ensure points is serializable
        }
        for apartment in apartments
    ]

    # Convert the list to JSON for frontend use
    houses_json = json.dumps(apartments_with_properties)

    # Render the template with context
    return render(
        request,
        'pagos/house_list.html',
        {
            'projects': projects,
            'apartments': apartments,
            'project_name': project_name,
            'houses_json': houses_json,
        }
    )

#------------------------------

import json
from django.shortcuts import render
from django.http import JsonResponse

def image_map_view(request):
    # Get all projects for the dropdown
    projects = Project.objects.all()

    # Get the selected project ID from the GET request
    selected_project_id = request.GET.get('project')

    # Initialize variables
    apartments = []
    houses_json = None
    project_name = "All Projects"
    project_plano = None

    if selected_project_id:
        # Validate and filter apartments for the selected project
        selected_project = get_object_or_404(Project, id=selected_project_id)
        project_name = selected_project.name
        project_plano = selected_project.plano.url
        apartments = Apartment.objects.filter(project=selected_project).order_by('number')

    # Create a list of apartments with necessary properties
    apartments_with_properties = [
        {
            "id": apartment.number,
            "status": apartment.status,
            "points": apartment.points,  # Ensure points is serializable
        }
        for apartment in apartments
    ]

    # Convert the list to JSON for frontend use
    houses_json = json.dumps(apartments_with_properties)

    # Render the template with context
    return render(
        request,
        'pagos/house_list2.html',
        {
            'project_plano': project_plano,
            'projects': projects,
            'apartments': apartments,
            'project_name': project_name,
            'houses_json': houses_json,
        }
    )


#------------------------------

def plan_edificio_view(request):
    building = []
    for piso in range(1, 9):  # Floors from 1 to 8
        floor = {
            'number': piso,
            'apartments': [((piso - 1) * 3 + depto) for depto in range(1, 4)]
        }
        building.append(floor)
    building.reverse()  # Reverse the order of floors so that Floor 1 is at the bottom
    return render(request, 'pagos/plan_edificio.html', {'building': building})


#------------------------------

from django.shortcuts import get_object_or_404, render
from .models import Project, Apartment, PriceList

def apartment_detail(request, project_name, apartment_number):
    # Retrieve the project by its name
    project = get_object_or_404(Project, name=project_name)

    # Retrieve the apartment by its number and project
    apartment = get_object_or_404(Apartment, project=project, number=apartment_number)

    # Calculate the total price of the apartment
    price_list = PriceList.objects.filter(apartment=apartment).first()
    total_price = None
    precio_por_m2 = None
    error_message = None

    if price_list:
        total_price = apartment.area * price_list.current_list_price
        precio_por_m2 = total_price / apartment.area
    else:
        error_message = "No price list available for this apartment."

    # Pass the apartment details and price info to the template
    context = {
        'apartment': apartment,
        'project': project,
        'total_price': total_price,
        'precio_por_m2': precio_por_m2,
        'error_message': error_message,
    }
    return render(request, 'pagos/apartment_detail.html', context)


#------------------------------
#### 3.3 Cotización ###
#------------------------------

def cotizacion(request):
    projects = Project.objects.all()
    apartments = None
    selected_apartment = None
    selected_project = None
    selected_price_list = None
    total_price = None

    if request.method == 'POST':
        project_id = request.POST.get('project')
        apartment_number = request.POST.get('apartment_number')

        if project_id:
            selected_project = get_object_or_404(Project, id=project_id)
            apartments = Apartment.objects.filter(project=selected_project).order_by('number')

            if apartment_number:
                selected_apartment = apartments.filter(number=apartment_number).first()

                # Fetch the PriceList for the selected apartment
                if selected_apartment:
                    selected_price_list = PriceList.objects.filter(apartment=selected_apartment).first()
                    total_price = selected_apartment.area * selected_price_list.current_list_price

    return render(request, 'pagos/cotizacion.html', {
        'projects': projects,
        'apartments': apartments,
        'selected_apartment': selected_apartment,
        'selected_project': selected_project,
        'selected_price_list': selected_price_list,
        'total_price': total_price,
    })


# -------------------------------

import calendar
from datetime import timedelta, date
from decimal import Decimal

def calcular_plan_pagos(precio_lista, porcentaje_enganche, porcentaje_descuento, porcentaje_mensualidades, num_mensualidades, mes_inicio):
    """Calculates the payment plan based on various parameters."""
    
    # Ensure precio_lista is a Decimal
    precio_lista = Decimal(precio_lista)

    # Calculate total percentages for payments
    porcentaje_pago_final = Decimal(100) - porcentaje_enganche - porcentaje_mensualidades 
    precio_descuento = precio_lista * (Decimal(1) - porcentaje_descuento / Decimal(100))
    
    # Calculate payments
    enganche = precio_descuento * (porcentaje_enganche / Decimal(100))
    pago_mensual = precio_descuento * (porcentaje_mensualidades / Decimal(100)) / Decimal(num_mensualidades)
    pago_final = precio_descuento * (porcentaje_pago_final / Decimal(100))
    
    # Generate payment schedule
    pagos = []
    current_date = mes_inicio

    # Add enganche payment
    pagos.append({
        'fecha': current_date,  # Use actual date
        'pago': float(enganche),
        'tipo': 'Enganche'
    })

    # Add monthly payments
    for _ in range(num_mensualidades):
        last_day = calendar.monthrange(current_date.year, current_date.month)[1]
        current_date += timedelta(days=last_day)  # Move to the next month
        pagos.append({
            'fecha': current_date,  # Use actual date
            'pago': float(pago_mensual),
            'tipo': 'Mensualidad'
        })

    # Add final payment
    last_day = calendar.monthrange(current_date.year, current_date.month)[1]
    current_date += timedelta(days=last_day)
    pagos.append({
        'fecha': current_date,  # Use actual date
        'pago': float(pago_final),
        'tipo': 'Pago final'
    })

    # Compile payment plan
    plan_pagos = {
        'pagos': pagos,
        'total': float(enganche + (pago_mensual * Decimal(num_mensualidades)) + pago_final),
    }

    return plan_pagos


# -------------------------------
from django.core.mail import send_mail
from django.conf import settings
from decimal import Decimal  # Ensure this is at the top of the file, not inside the function

def payment_plan_view(request):
    total_price = request.GET.get('total_price')
    apartment_number = request.GET.get('apartment_number')
    selected_project_name = request.GET.get('selected_project')

    # Convert total_price to Decimal, handling None case
    total_price = Decimal(total_price) if total_price else Decimal(0)

    apartment = Apartment.objects.filter(number=apartment_number).first()
    project = Project.objects.filter(name=selected_project_name).first()

    if not apartment or not project:
        return render(request, 'pagos/plan_pagos.html', {
            'error': 'Invalid apartment or project selected.'
        })

    form = PaymentPlanForm(request.POST or None)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'calculate' and form.is_valid():
            data = form.cleaned_data
            mes_inicio = data['mes_inicio']
            plan_pagos = calcular_plan_pagos(
                total_price,
                data['porcentaje_enganche'],
                data['porcentaje_descuento'],
                data['porcentaje_mensualidades'],
                data['num_mensualidades'],
                mes_inicio
            )

            pagos = plan_pagos.get('pagos', [])
            npv = calculate_npv(pagos, 0.05)
            print (npv)

            """npv = Decimal(npv) if isinstance(npv, (float, str)) else Decimal(0)
            ratio = npv / total_price if npv > 0 else None"""

            ratio = 1
            indicator_color = (
                'green' if ratio and ratio >= 0.95 else
                'yellow' if ratio and ratio >= 0.90 else
                'red'
            )

            return render(request, 'pagos/plan_pagos.html', {
                'form': form,
                'plan_pagos': plan_pagos,
                'npv': npv,
                'indicator_color': indicator_color,
                'ratio': ratio,
            })

        elif action == 'cotizacion' and form.is_valid():
            data = form.cleaned_data
            mes_inicio = data['mes_inicio']
            plan_pagos = calcular_plan_pagos(
                total_price,
                data['porcentaje_enganche'],
                data['porcentaje_descuento'],
                data['porcentaje_mensualidades'],
                data['num_mensualidades'],
                mes_inicio
            )

            cliente = data.get('cliente')
            if not cliente:
                return render(request, 'pagos/plan_pagos.html', {
                    'form': form,
                    'error': 'No se seleccionó cliente.'
                })

            cliente_email = cliente.get_email()
            pagos = plan_pagos.get('pagos', [])
            total = plan_pagos.get('total', Decimal(0))

            subject = f"Cotización del {selected_project_name} - Apartamento {apartment_number}"
            email_message = f"""
Estimado cliente,

Aquí está su cotización para el proyecto **{selected_project_name}** y el apartamento **{apartment_number}** con precio total: ${total_price:,.2f}.

Resultados del Plan de Pagos:
"""

            for pago in pagos:
                fecha = pago.get('fecha', 'Fecha no disponible')
                tipo = pago.get('tipo', 'Tipo no disponible')
                pago_amount = pago.get('pago', Decimal(0))  # Default to Decimal(0) if missing
                email_message += f"Fecha {fecha} - {tipo} - ${pago_amount:.2f}\n"

            email_message += f"\nTotal a Pagar: ${total:.2f}\n"
            email_message += "\nAtentamente,\nEl equipo de ventas"

            send_mail(
                subject=subject,
                message=email_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[cliente_email],
                fail_silently=False,
            )

            return render(request, 'pagos/admin_venta_menu.html', {
                'message': 'Cotización enviada al cliente exitosamente.'
            })

    return render(request, 'pagos/plan_pagos.html', {
        'form': form,
        'total_price': total_price,
        'apartment_number': apartment_number,
        'selected_project': selected_project_name,
    })


#------------------------------------
#### 3.4 Proceso de venta o apartado ###
#------------------------------------
from django.db.models import Prefetch
from django.shortcuts import render

def apartado_venta(request):
    # Fetch all projects for the filter dropdown
    all_projects = Project.objects.all()

    # Get filter and sorting values from query parameters
    project_name_filter = request.GET.get('project_name', '')
    order_by = request.GET.get('order_by', 'number')  # Default to 'number'

    # Validate order_by to prevent invalid field names
    valid_order_fields = ['number', 'status']  # Define allowed fields
    if order_by not in valid_order_fields:
        order_by = 'number'

    # Prefetch apartments with the desired ordering
    apartments_prefetch = Prefetch(
        'apartments',
        queryset=Apartment.objects.order_by(order_by)
    )

    # Filter projects if a project name is provided
    if project_name_filter:
        projects = Project.objects.prefetch_related(apartments_prefetch).filter(name=project_name_filter)
    else:
        projects = Project.objects.prefetch_related(apartments_prefetch)

    return render(request, 'pagos/venta_apartado.html', {
        'projects': projects,
        'all_projects': all_projects,
        'project_name_filter': project_name_filter,
    })


#------------------------------------

def apartar(request, apartment_id):
    # Fetch the apartment, or return 404 if it doesn't exist
    apartment = get_object_or_404(Apartment, id=apartment_id)

    if request.method == 'POST':
        # Update apartment status to "apartado" (reserved)
        apartment.status = 'apartado'
        apartment.save()  # Save the updated status to the database

        # Redirect to the list of apartments or an appropriate view
        return redirect('pagos:lista_departamentos')  # Make sure this URL name exists in urls.py

    # Render the confirmation page for reservation
    return render(request, 'pagos/apartar.html', {
        'apartment': apartment,
    })
# -------------------------------

def disponible_view(request,apartment_id):
    # Fetch the apartment, or return 404 if it doesn't exist
    apartment = get_object_or_404(Apartment, id=apartment_id)

    if request.method == 'POST':
        # Update apartment status to "apartado" (reserved)
        apartment.status = 'disponible'
        apartment.save()  # Save the updated status to the database
        Venta.objects.filter(apartment=apartment).delete()
        PaymentRecord.objects.filter(apartment=apartment).delete()

        # Redirect to the list of apartments or an appropriate view
        return redirect('pagos:lista_departamentos')  # Make sure this URL name exists in urls.py

    # Render the confirmation page for reservation
    return render(request, 'pagos/disponible.html', {
        'apartment': apartment,
})

# -------------------------------

from django.shortcuts import get_object_or_404, redirect, render
from .models import PaymentRecord, PaymentInstallment
from .forms import VentaPlanForm
from clientes.models import Interaction

def ventas(request, apartment_id):
    # Fetch the apartment and its related project
    apartment = get_object_or_404(Apartment, id=apartment_id)
    project = apartment.project
    plan_pagos = None
    error_message = None
    total_price = None

    # Create an instance of PaymentPlanForm
    venta_plan_form = VentaPlanForm(request.POST or None, project=project)

    # Fetch the price list and calculate the total price
    price_list = PriceList.objects.filter(apartment=apartment).first()
    if price_list:
        total_price = apartment.area * price_list.current_list_price
    else:
        error_message = "No price list available for this apartment."

    # Handle form submission
    if request.method == 'POST' and venta_plan_form.is_valid():
        oportunidad = venta_plan_form.cleaned_data['cliente']
        porcentaje_descuento = venta_plan_form.cleaned_data['porcentaje_descuento']
        porcentaje_enganche = venta_plan_form.cleaned_data['porcentaje_enganche']
        porcentaje_mensualidades = venta_plan_form.cleaned_data['porcentaje_mensualidades']
        num_mensualidades = venta_plan_form.cleaned_data['num_mensualidades']
        mes_inicio = venta_plan_form.cleaned_data['mes_inicio']

        if total_price:  # Proceed only if total_price is valid
            try:
                # Calculate payment plan
                plan_pagos = calcular_plan_pagos(
                    total_price,
                    porcentaje_enganche,
                    porcentaje_descuento,
                    porcentaje_mensualidades,
                    num_mensualidades,
                    mes_inicio,
                )
                print (plan_pagos)

                # Save the payment record (general details)
                payment_record = PaymentRecord.objects.create(
                    oportunidad=oportunidad,
                    apartment=apartment,
                    project=project,
                    porcentaje_descuento=porcentaje_descuento,
                    porcentaje_enganche=porcentaje_enganche,
                    porcentaje_mensualidades=porcentaje_mensualidades,
                    num_mensualidades=num_mensualidades,
                    mes_inicio=mes_inicio,
                )

                # Create individual PaymentInstallment objects for each mensualidad
                for num, pago in enumerate(plan_pagos['pagos'], start=1):
                    PaymentInstallment.objects.create(
                        payment_record=payment_record,
                        installment_number=num,  # Update to match the model field name
                        due_date=pago['fecha'],  # Update to match the model field name
                        total_amount=pago['pago'],  # Update to match the model field name
                        amount_paid=0.00,  # Update to match the model field name
                        fully_paid=False,  # Update to match the model field name
                    )
                
                # Register the sale in the Ventas model
                cliente = oportunidad.cliente
                
                Venta.objects.create(
                    project=project,
                    apartment=apartment,
                    cliente=cliente,
                    fecha_venta=timezone.now()  # Use the current date as the sale date
                )

                Interaction.objects.create(
                    cliente=cliente,
                    oportunidad=oportunidad,
                    salesperson=request.user,  # Assuming the salesperson is the current logged-in user
                    interaction_type='Other',
                    category='Venta',
                    notes='Sale registered for the apartment.'  # You can customize the notes if needed
                )


                # Update apartment status
                apartment.status = "vendido"
                apartment.save()
                oportunidad.estatus = "vendido"
                oportunidad.save()

                return redirect('pagos:list_payment_records')

            except ValueError as e:
                error_message = f"Error in payment calculation: {str(e)}"

    # Render the template
    return render(request, 'pagos/venta.html', {
        'apartment': apartment,
        'project': project,
        'plan_pagos': plan_pagos,
        'error_message': error_message,
        'venta_plan_form': venta_plan_form,
    })


# -------------------------------

def save_payment_record(request, form, apartment, project, total_price):
    data = form.cleaned_data
    cliente = data['cliente']
    porcentaje_descuento = data['porcentaje_descuento']
    porcentaje_enganche = data['porcentaje_enganche']
    porcentaje_mensualidades = data['porcentaje_mensualidades']
    num_mensualidades = data['num_mensualidades']
    mes_inicio = data['mes_inicio']

    # Calculate amounts
    discounted_price = total_price * (1 - porcentaje_descuento / 100)
    down_payment_amount = discounted_price * (porcentaje_enganche / 100)
    monthly_payment_amount = (discounted_price - down_payment_amount) / num_mensualidades

    # Create PaymentRecord
    payment_record = PaymentRecord.objects.create(
        cliente=cliente,
        apartment=apartment,
        project=project,
        porcentaje_descuento=porcentaje_descuento,
        porcentaje_enganche=porcentaje_enganche,
        porcentaje_mensualidades=porcentaje_mensualidades,
        num_mensualidades=num_mensualidades,
        mes_inicio=mes_inicio,
    )

    # Generate PaymentInstallments
    installments = []
    for i in range(1, num_mensualidades + 1):
        due_date = mes_inicio + timedelta(days=30 * i)  # Adjust for each month's due date
        installment = PaymentInstallment(
            payment_record=payment_record,
            installment_number=i,
            due_date=due_date,
            total_amount=monthly_payment_amount,
            amount_paid=0.00,
            fully_paid=False,
        )
        installments.append(installment)

    PaymentInstallment.objects.bulk_create(installments)  # Save all installments at once

    return True


# -------------------------------

def pay_installment(installment_id, payment_amount):
    try:
        installment = PaymentInstallment.objects.get(id=installment_id)
        installment.amount_paid += payment_amount
        installment.update_payment_status()
        return f"Installment {installment.installment_number} updated successfully."
    except PaymentInstallment.DoesNotExist:
        return "Installment not found."


# -------------------------------
#### 3.6 Pagos y estado de cuenta $$$
# -------------------------------

from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from pagos.models import PaymentRecord, Cliente  # Adjust imports as needed

@login_required
def list_payment_records(request):
    # Check if the user is an admin or superuser
    is_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    is_cliente = request.user.groups.filter(name='Cliente').exists()

    # Initialize `records` to avoid issues if no condition matches
    records = PaymentRecord.objects.none()

    if is_admin:
        # Admins can view all payment records
        records = PaymentRecord.objects.all()
    elif is_cliente:
        # Fetch the Cliente profile associated with the logged-in user
        cliente = request.user.userprofile
        print(cliente)
        # Retrieve payment records for this Cliente
        records = PaymentRecord.objects.filter(cliente=cliente)

    # Render the template with the retrieved records and admin flag
    return render(request, 'pagos/payment_records_list.html', {'records': records, 'is_admin': is_admin, 'is_cliente': is_cliente})


# -------------------------------

from django.db.models import Sum
from django.shortcuts import get_object_or_404, render

def record_detail_view(request, pk):
    # Fetch the payment record based on the primary key
    record = get_object_or_404(PaymentRecord, pk=pk)
    
    # Fetch the associated payment installments (schedule)
    payment_schedule = PaymentInstallment.objects.filter(payment_record=record)
    
    # Calculate total amount due (sum of total_amount for all installments)
    total_amount_due = payment_schedule.aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # Calculate total amount paid (sum of amount_paid for all installments)
    total_amount_paid = payment_schedule.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

    # Calculate remaining balance
    remaining_balance = total_amount_due - total_amount_paid

    # Render the template with the payment record, payment schedule, and calculated totals
    return render(request, 'pagos/payment_record_detail.html', {
        'record': record,
        'payment_schedule': payment_schedule,
        'total_amount_due': total_amount_due,
        'total_amount_paid': total_amount_paid,
        'remaining_balance': remaining_balance,
    })

# -------------------------------

def delete_record(request, id):
    record = get_object_or_404(PaymentRecord, id=id)
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'Record deleted successfully.')
    return redirect('pagos:list_payment_records')  # Replace with the name of your listing view

# -------------------------------


from django.shortcuts import get_object_or_404, redirect
from pagos.models import PaymentInstallment

def cancel_payment(request, installment_id):
    # Retrieve the PaymentInstallment by its ID
    payment = get_object_or_404(PaymentInstallment, id=installment_id)

    # Check if the request is POST (to prevent accidental GET calls)
    if request.method == 'POST':
        # Mark as not fully paid and reset the amount paid
        if payment.fully_paid:
            payment.amount_paid = 0  # Reset payment amount
            payment.fully_paid = False
            payment.save()

        # Redirect to the related PaymentRecord detail page
        return redirect('pagos:record_detail', pk=payment.payment_record.id)

    # If not POST, redirect back with no changes
    return redirect('pagos:record_detail', pk=payment.payment_record.id)


# -------------------------------
from django.shortcuts import get_object_or_404, redirect
from pagos.models import PaymentInstallment

def toggle_payment_status(request, installment_id):
    # Retrieve the PaymentInstallment by its ID
    payment = get_object_or_404(PaymentInstallment, id=installment_id)

    # Check if the request is POST (to prevent accidental GET calls)
    if request.method == 'POST':
        # Mark as fully paid if not already paid
        if not payment.fully_paid:
            payment.amount_paid = payment.total_amount  # Ensure full payment
            payment.fully_paid = True
            payment.save()

        # Redirect to the related PaymentRecord detail page using the correct parameter name
        return redirect('pagos:record_detail', pk=payment.payment_record.id)

    # If not POST, redirect back with no changes
    return redirect('pagos:record_detail', pk=payment.payment_record.id)

# -----------------------------

from django.shortcuts import get_object_or_404, render, redirect
from datetime import datetime
from bancos.models import Transaction, BankAccount  # Import the relevant models
from pagos.models import PaymentRecord  # Assuming PaymentRecord is your payment model

def register_payment(request, payment_record_id):
    # Retrieve the PaymentRecord using its ID
    payment_record = get_object_or_404(PaymentRecord, id=payment_record_id)

    # Retrieve the associated Project for the PaymentRecord
    project = payment_record.project

    if request.method == 'POST':
        # Retrieve the amount paid from the form submission
        amount_paid = request.POST.get('amount_paid')
        amount_paid = Decimal(request.POST.get('amount_paid'))

        # Create the PaymentTransaction (this will be your payment record)
        payment_transaction = PaymentTransaction.objects.create(
            payment_record=payment_record,
            amount_paid=amount_paid,
            payment_date=datetime.today().date(),
        )

        # Retrieve the associated BankAccount from the Project
        
        bank_account = project.bank_accounts.first()  # Assuming there's only one BankAccount per Project
        
        if bank_account:  # Check if the BankAccount exists
            # Create a Deposit Transaction in the bancos app
            bank_account.deposit(amount_paid, description=f"Payment for {payment_record.cliente}'s apartment {payment_record.apartment}")
                        
            Transaction.objects.create(
                bank_account=bank_account,
                amount=amount_paid,
                transaction_type=Transaction.DEPOSIT,  # Since the payment is a deposit
                description=f"Payment for {payment_record.cliente}'s apartment {payment_record.apartment}"
            )
        else:
            # Handle the case where no BankAccount is found for the Project
            print("No BankAccount found for the project.")

        # Redirect to the list of payment records after successful payment registration
        return redirect('pagos:list_payment_records')

    # If it's a GET request, render the form to register the payment
    return render(request, 'pagos/register_payment.html', {
        'payment_record': payment_record,
    })

# -------------------------------
#### 3.6 Ventas ###
# -------------------------------

def review_ventas(request):
    # Fetch all Ventas objects, optionally ordered by fecha_venta
    ventas = Venta.objects.all().order_by('-fecha_venta')  # Order by most recent sale

    # Render the ventas in the template
    return render(request, 'pagos/review_ventas.html', {'ventas': ventas})





# -------------------------------
#### Otros ###
# -------------------------------

def calculate_npv(payment_schedule, discount_rate):
    # Convert discount rate to a monthly rate
    monthly_discount_rate = (1 + discount_rate) ** (1/12) - 1
    
    # Determine the start date for calculating periods
    try:
        start_date = payment_schedule[0].get('fecha')
        if not isinstance(start_date, datetime.date):
            raise ValueError("Start date is not a valid datetime.date object")
    except (IndexError, ValueError, TypeError):
        return "Error: 'fecha' date format issue in payment schedule."
    
    npv = 0.0
    
    for payment_entry in payment_schedule:
        # Safely get values from the dictionary with defaults if missing
        payment_date = payment_entry.get('fecha')
        payment_amount = payment_entry.get('pago', 0)
        
        # Skip entries where 'fecha' or 'pago' is missing
        if payment_date is None or payment_amount is None:
            continue
        
        # Calculate the period in months from the start date
        period = (payment_date.year - start_date.year) * 12 + (payment_date.month - start_date.month)
        
        # Add discounted payment to NPV
        npv += payment_amount / ((1 + monthly_discount_rate) ** period)
    
    return npv





# -------------------------------
#### Canva ###
# -------------------------------


from django.shortcuts import render
from .models import House
from .forms import HouseForm
import json

def plano(request):
    houses = House.objects.all()  # Get all houses
    print(houses)

    # Create a JSON-friendly list with calculated properties
    houses_with_properties = [
        {
            "id": house.id,
            "name": house.name,
            "color": house.color,
            "points": house.points,
            "area": house.calculate_area(),
            "width": house.calculate_dimensions()[0],
            "height": house.calculate_dimensions()[1]
        }
        for house in houses
    ]
    
    # Convert the list to JSON for frontend use
    houses_json = json.dumps(houses_with_properties)

    # Pass both houses_with_properties and houses_json to the template
    return render(
        request,
        "pagos/house_list.html",
        {"houses": houses_with_properties, "houses_json": houses_json}
    )




# views.py
import json
from openpyxl import load_workbook
from django.shortcuts import render
from django.http import HttpResponse
from .forms import CSVUploadForm  # You can rename this form to be more appropriate, like 'ExcelUploadForm'
from .models import House

def upload_excel(request):
    if request.method == "POST" and request.FILES["excel_file"]:
        excel_file = request.FILES["excel_file"]

        # Load the Excel workbook
        wb = load_workbook(excel_file)
        sheet = wb.active  # Assuming the data is in the active sheet

        # Iterate over the rows in the sheet (skipping header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]  # Name of the house
            color = row[1]  # Color of the house
            points_str = row[2]  # Points as a string

            try:
                # Safely load points as a list of coordinates using json.loads
                points = json.loads(points_str)
            except json.JSONDecodeError:
                return HttpResponse(f"Error: Invalid points format in row {row}")

            # Create a new house instance
            house = House(name=name, color=color, points=points)
            house.save()

        return HttpResponse("Excel file processed successfully!")

    form = CSVUploadForm()  # Adjust the form if needed
    return render(request, "pagos/upload_excel.html", {"form": form})


#-----------------

# views.py

# views.py

import pandas as pd
from django.shortcuts import render
from django.http import HttpResponse
from .models import PaymentInstallment, Project

def generar_reporte_ingresos(request):
    # Capturar los valores del filtro desde los parámetros GET
    project_name_filter = request.GET.get('project_name', '')
    
    # Filtrar las cuotas con base en el nombre del proyecto
    cuotas = PaymentInstallment.objects.all()
    
    if project_name_filter:
        cuotas = cuotas.filter(payment_record__project__name=project_name_filter)
    
    # Extraer los datos necesarios
    data = cuotas.values(
        'due_date',  # Fecha de vencimiento
        'total_amount',  # Monto total de la cuota
        'amount_paid',  # Monto pagado hasta la fecha
        'payment_record__project__name',  # Nombre del proyecto
        'payment_record__apartment__number',  # Número del apartamento
    )
    
    # Convertir los datos en un DataFrame para procesar
    df = pd.DataFrame(data)
    
    # Convertir la fecha de vencimiento al primer día del mes correspondiente
    df['mes'] = pd.to_datetime(df['due_date'], errors='coerce').dt.to_period('M').dt.to_timestamp()

    # Crear una columna de ingresos reales (total pagado)
    df['ingresos_reales'] = df['amount_paid'].fillna(0)  # Usamos 0 si no hay pago

    # Crear una columna de ingresos proyectados (total por pagar)
    df['ingresos_proyectados'] = df['total_amount']

    # Agrupar por mes y proyecto
    ingresos_mensuales = df.groupby(['mes', 'payment_record__project__name']).agg(
        total_ingresos_reales=('ingresos_reales', 'sum'),
        total_ingresos_proyectados=('ingresos_proyectados', 'sum')
    ).reset_index()

    # Convertir los DataFrames a listas de diccionarios para usar en el template
    ingresos_mensuales_list = ingresos_mensuales.to_dict('records')

    # Obtener todos los proyectos para llenar el dropdown
    all_projects = Project.objects.all()

    # Si es una solicitud de exportación, devolvemos el archivo Excel
    if request.GET.get('export') == 'true':
        with pd.ExcelWriter('reporte_ingresos_completo.xlsx') as writer:
            ingresos_mensuales.to_excel(writer, sheet_name='Ingresos por Proyecto', index=False)

        with open('reporte_ingresos_completo.xlsx', 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="reporte_ingresos_completo.xlsx"'
            return response

    # Pasamos los datos al template para mostrar en la página web
    context = {
        'ingresos_mensuales': ingresos_mensuales_list,
        'all_projects': all_projects,  # Para el dropdown de proyectos
        'project_name_filter': project_name_filter,  # Para mantener el filtro seleccionado
    }
    return render(request, 'pagos/reporte_ingresos.html', context)


